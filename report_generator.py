# report_generator.py

import pandas as pd
import os
import unicodedata
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import config
from fpdf import FPDF

# --- FUNÇÕES AUXILIARES E DE FORMATAÇÃO (sem alterações) ---

def formatar_nome_de_email(email, apenas_primeiro_nome=False):
    """Extrai um nome formatado de um e-mail, com opção para retornar apenas o primeiro nome."""
    if not isinstance(email, str) or '@' not in email:
        return "N/A"
    try:
        nome_completo = email.split('@')[0].replace('.', ' ').title()
        if apenas_primeiro_nome:
            return nome_completo.split(' ')[0]
        return nome_completo
    except:
        return email

def sanitizar_nome_arquivo(nome):
    nfkd_form = unicodedata.normalize('NFKD', nome)
    nome_ascii = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    nome_seguro = re.sub(r'[^\w\s-]', '', nome_ascii).strip().replace(' ', '_')
    return nome_seguro

def criar_nome_arquivo_loja(nome_loja):
    """Cria um nome de arquivo legível, removendo apenas caracteres inválidos."""
    # Remove caracteres que não são permitidos em nomes de arquivo
    nome_seguro = re.sub(r'[\\/*?:"<>|]', "", nome_loja)
    return f"Relatório de Rupturas {nome_seguro}.xlsx"

def get_numero_loja(nome_loja):
    match = re.match(r"^\d+", nome_loja)
    return int(match.group(0)) if match else None

def formatar_excel(caminho_arquivo):
    print(f"Formatando arquivo: {caminho_arquivo}...")
    try:
        workbook = load_workbook(caminho_arquivo)
        header_fill = PatternFill(start_color="e60d25", end_color="e60d25", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_index, row in enumerate(sheet.iter_rows(), 1):
                if row_index == 1:
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                elif row_index % 2 == 0:
                    for cell in row:
                        cell.fill = even_row_fill
            
            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(caminho_arquivo)
        print("Formatação e ajuste de colunas aplicados com sucesso.")
    except Exception as e:
        print(f"Não foi possível formatar o Excel. Erro: {e}")

# --- FUNÇÃO DE RELATÓRIO DOS GERENTES (COM SAUDAÇÃO PERSONALIZADA) ---

def gerar_relatorios_gerentes(creds, df, pasta_destino, enviar_email_func):
    print("\n--- INICIANDO GERAÇÃO DE RELATÓRIOS POR LOJA ---")

    for loja, email_gerente in config.GERENTES_EMAILS.items():
        print(f"\nProcessando loja: {loja}...")
        df_loja = df[df["loja"] == loja].copy()
        
        if df_loja.empty:
            print(f"Nenhum dado encontrado para a loja {loja}. Pulando.")
            continue

        categorias_validas = [cat for cat in df_loja["categoria"].unique() if cat and not pd.isna(cat)]

        if not categorias_validas:
            print(f"Nenhuma categoria válida encontrada para os registros da loja {loja}. Pulando.")
            continue
            
        destinatario = config.EMAIL_TESTE if config.MODO_TESTE else email_gerente
        nome_gerente = formatar_nome_de_email(email_gerente, apenas_primeiro_nome=True)
        
        total_rupturas = len(df_loja)
        
        # --- INÍCIO DA CORREÇÃO ---
        # Converte strings vazias para nulo, depois preenche.
        df_loja['tratativa'] = df_loja['tratativa'].replace('', pd.NA)
        df_loja["tratativa"] = df_loja["tratativa"].fillna("Sem Tratativa")
        # --- FIM DA CORREÇÃO ---
        
        rupturas_tratadas = len(df_loja[df_loja["tratativa"] != "Sem Tratativa"])
        divergencias = df_loja[df_loja["tratativa"] == "Verificar Estoque (Divergência)"]
        
        lista_divergencias_html = "<li>Nenhuma divergência apontada.</li>"
        if not divergencias.empty:
            lista_divergencias_html = "".join([f"<li>{row.produto} (Cód: {row.codigo_produto})</li>" for _, row in divergencias.iterrows()])
        
        corpo_email = f'<html><body><h2>Relatório de Rupturas - {loja}</h2><p>Olá, {nome_gerente},</p><p>Segue o resumo das rupturas identificadas em sua loja:</p><ul><li><b>Total de Rupturas Identificadas:</b> {total_rupturas}</li><li><b>Rupturas com Tratativa:</b> {rupturas_tratadas}</li></ul><hr><h3>Produtos com Tratativa "Verificar Estoque (Divergência)":</h3><ul>{lista_divergencias_html}</ul><hr><p>O relatório completo, com todas as rupturas separadas por categoria, está em anexo.</p><p>Atenciosamente,<br>Equipe Comercial</p></body></html>'
        
        nome_arquivo = criar_nome_arquivo_loja(loja)
        caminho_completo_arquivo = os.path.join(pasta_destino, nome_arquivo)
        
        with pd.ExcelWriter(caminho_completo_arquivo, engine='openpyxl') as writer:
            colunas_relatorio = ["timestamp", "nome_solicitante", "categoria", "produto", "tempo_ruptura", "tratativa"]
            colunas_rename = {"timestamp": "Data Solicitação", "nome_solicitante": "Solicitante", "categoria": "Categoria", "produto": "Produto", "tempo_ruptura": "Tempo de Ruptura", "tratativa": "Tratativa"}
            
            for categoria in categorias_validas:
                df_categoria = df_loja[df_loja["categoria"] == categoria]
                df_final = df_categoria[colunas_relatorio].rename(columns=colunas_rename)
                df_final.to_excel(writer, sheet_name=sanitizar_nome_arquivo(str(categoria))[:31], index=False)
        
        print(f"Arquivo Excel '{caminho_completo_arquivo}' gerado.")
        formatar_excel(caminho_completo_arquivo)
        enviar_email_func(creds, destinatario, f"Relatório de Rupturas - {loja}", corpo_email, caminho_completo_arquivo)

# --- FUNÇÃO DE RELATÓRIO DOS COMPRADORES (COM SAUDAÇÃO E CORPO ATUALIZADOS) ---

def gerar_relatorios_compradores(creds, df, pasta_destino, enviar_email_func):
    print("\n--- INICIANDO GERAÇÃO DE ALERTAS PARA COMPRADORES ---")
    df_pedidos = df[df["tratativa"] == "Será feito pedido"].copy()
    if df_pedidos.empty:
        print("Nenhuma ruptura com tratativa 'Será feito pedido' encontrada.")
        return

    pedidos_por_categoria = df_pedidos.groupby('categoria')

    for categoria, df_categoria in pedidos_por_categoria:
        print(f"\nProcessando categoria para compradores: {categoria}...")
        primeira_linha = df_categoria.iloc[0]
        num_loja_exemplo = get_numero_loja(primeira_linha['loja'])
        email_comprador = None
        if not num_loja_exemplo: continue

        if num_loja_exemplo in config.LOJAS_PB:
            email_comprador = config.COMPRADORES_PB_EMAILS.get(categoria)
        elif num_loja_exemplo in config.LOJAS_RN1 or num_loja_exemplo in config.LOJAS_RN2:
            if categoria == "Bebidas":
                email_comprador = config.COMPRADORES_RN_BEBIDAS.get("RN1") if num_loja_exemplo in config.LOJAS_RN1 else config.COMPRADORES_RN_BEBIDAS.get("RN2")
            else:
                email_comprador = config.COMPRADORES_RN_EMAILS.get(categoria)

        if not email_comprador:
            print(f"Nenhum comprador encontrado para a categoria '{categoria}'. Pulando.")
            continue

        destinatario = config.EMAIL_TESTE if config.MODO_TESTE else email_comprador
        
        # *** NOVA LÓGICA: Extrai o primeiro nome do comprador para a saudação ***
        nome_comprador = formatar_nome_de_email(email_comprador, apenas_primeiro_nome=True)
        
        nome_arquivo = f"Relatorio_Compras_{sanitizar_nome_arquivo(categoria)}.xlsx"
        caminho_completo_arquivo = os.path.join(pasta_destino, nome_arquivo)

        with pd.ExcelWriter(caminho_completo_arquivo, engine='openpyxl') as writer:
            pedidos_por_loja = df_categoria.groupby('loja')
            for loja, df_loja in pedidos_por_loja:
                colunas_relatorio = ["codigo_produto", "produto", "nome_solicitante", "timestamp"]
                colunas_rename = {"codigo_produto": "Código", "produto": "Produto", "nome_solicitante": "Solicitante", "timestamp": "Data Solicitação"}
                df_loja_final = df_loja[colunas_relatorio].rename(columns=colunas_rename)
                df_loja_final.to_excel(writer, sheet_name=sanitizar_nome_arquivo(loja)[:31], index=False)
        
        print(f"Arquivo Excel '{caminho_completo_arquivo}' gerado para o comprador de '{categoria}'.")
        formatar_excel(caminho_completo_arquivo)

        # *** ATUALIZADO: Corpo do e-mail com saudação personalizada e nome da categoria ***
        corpo_email = f'<html><body><h2>Alerta de Pedido de Compra - Categoria: {categoria}</h2><p>Olá, {nome_comprador},</p><p>Segue em anexo a lista de produtos da categoria <b>{categoria}</b> que precisam de pedido de compra, separados por loja.</p><br><p>Atenciosamente,<br>Equipe Comercial</p></body></html>'
        
        enviar_email_func(creds, destinatario, f"Alerta de Compra - {categoria}", corpo_email, caminho_completo_arquivo)

def gerar_relatorio_gerencial_pdf(df, pasta_destino, data_inicio, data_fim):
    print("\n--- INICIANDO GERAÇÃO DO RELATÓRIO GERENCIAL PDF ---")
    if df.empty:
        print("DataFrame vazio. Não é possível gerar o relatório gerencial.")
        return

    # 1. Agregação de Dados
    total_solicitacoes = len(df)
    
    # --- INÍCIO DA CORREÇÃO ---
    # Converte strings vazias para nulo, depois preenche.
    df['tratativa'] = df['tratativa'].replace('', pd.NA)
    df["tratativa"] = df["tratativa"].fillna("Sem Tratativa")
    # --- FIM DA CORREÇÃO ---
    
    # Análise por Loja
    solicitacoes_por_loja = df.groupby('loja').size().reset_index(name='total')
    tratativas_por_loja = df[df['tratativa'] != 'Sem Tratativa'].groupby('loja').size().reset_index(name='tratadas')
    
    # Juntando os dados de loja
    df_lojas = pd.merge(solicitacoes_por_loja, tratativas_por_loja, on='loja', how='left').fillna(0)
    df_lojas['tratadas'] = df_lojas['tratadas'].astype(int)
    
    # Análise por tipo de Tratativa
    resumo_tratativas = df['tratativa'].value_counts().reset_index()
    resumo_tratativas.columns = ['tratativa', 'quantidade']

    # 2. Geração do PDF
    # (O restante da função de gerar o PDF continua o mesmo)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    
    pdf.cell(0, 10, "Relatório Gerencial de Rupturas", 0, 1, "C")
    pdf.set_font("Arial", "", 12)
    periodo_str = f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    pdf.cell(0, 10, periodo_str, 0, 1, "C")
    pdf.ln(10)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Total de Solicitações no Período: {total_solicitacoes}", 0, 1)
    pdf.ln(5)

    pdf.cell(0, 10, "Desempenho por Loja:", 0, 1)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(80, 8, "Loja", 1, 0, "C")
    pdf.cell(40, 8, "Solicitações", 1, 0, "C")
    pdf.cell(40, 8, "Tratadas", 1, 1, "C")
    pdf.set_font("Arial", "", 10)
    for index, row in df_lojas.iterrows():
        pdf.cell(80, 8, row['loja'], 1)
        pdf.cell(40, 8, str(row['total']), 1, 0, "C")
        pdf.cell(40, 8, str(row['tratadas']), 1, 1, "C")
    pdf.ln(10)
    
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Resumo por Tipo de Tratativa:", 0, 1)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(120, 8, "Tratativa", 1, 0, "C")
    pdf.cell(40, 8, "Quantidade", 1, 1, "C")
    pdf.set_font("Arial", "", 10)
    for index, row in resumo_tratativas.iterrows():
        pdf.cell(120, 8, row['tratativa'], 1)
        pdf.cell(40, 8, str(row['quantidade']), 1, 1, "C")
    
    nome_arquivo = f"Relatorio_Gerencial_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.pdf"
    caminho_completo = os.path.join(pasta_destino, nome_arquivo)
    pdf.output(caminho_completo)
    print(f"Relatório Gerencial PDF '{caminho_completo}' gerado com sucesso.")
    
    return caminho_completo